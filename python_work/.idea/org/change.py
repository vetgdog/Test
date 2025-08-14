from tkinter.font import names

import openpyxl
from openpyxl import load_workbook


def write_to_excel(file_path, sheet_name, row, column, value):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 获取指定的sheet
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        raise ValueError(f"Sheet '{sheet_name}' 不存在")

    # 写入内容
    sheet.cell(row=row, column=column, value=value)

    # 保存Excel
    workbook.save(file_path)
    print(f"成功写入 '{value}' 到 {sheet_name}! 位置：第{row}行 第{column}列")



def build_intro(name, text):
    return text.replace('$$$$', name)




# names = [
#     "Luna",
#     "linling",
#     "Ayumi",
#     "Neytiri",
#     "Lara",
#     "Selina",
#     "Elsa",
#     "Natasha",
#     "Diana",
#     "Mia",
#     "Ariana Marie",
#     "Nagano ichika",
#     "Briar Riley",
#     "Sophie"
# ]
# names = [
#     "King Trumpy",
#     "Christopher"
# ]
names = [
    "Cleopatra",
    "Ariel"
]
# names = [
#     "Thor",
#     "Steve Rogers"
# ]





def read_texts_from_excel(file_path, sheet_name='Sheet1', column=1, start_row=2):
    """
    从 Excel 指定列读取文本数据，填入 texts 列表。

    :param file_path: Excel 文件路径
    :param sheet_name: 要读取的 Sheet 名称
    :param column: 要读取的列编号（从 1 开始）
    :param start_row: 从第几行开始读取（默认跳过表头，从第2行开始）
    :return: texts 列表
    """
    texts = []
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    # 逐行读取指定列的数据
    for row in sheet.iter_rows(min_row=start_row, min_col=column, max_col=column):
        cell = row[0]
        if cell.value is not None:
            texts.append(str(cell.value).strip())

    return texts

texts = read_texts_from_excel(r'C:\Users\admin\Desktop\工作簿change.xlsx',sheet_name='Sheet1', column=1, start_row=2)


j = 0;
for text in texts:
    for i in range(1,3):
        text1 = build_intro(names[i-1],text)
        write_to_excel(
                    file_path=r'C:\Users\admin\Desktop\工作簿texts.xlsx',
                    sheet_name='Sheet1',
                    row= 2+j,
                    column=1,
                    value=text1,
                    # value="Oh, baby, I’m so good at listening to you flirt!"
                )
        j = j + 1







# i = 2
# names = "Christopher"
#
# for text in texts:
#         text = build_intro(names,text)
#         write_to_excel(
#             file_path=r'C:\Users\admin\Desktop\工作簿_X2_2.xlsx',
#             sheet_name='Sheet1',
#             row= i,
#             column=1,
#             value=text,
#             # value="Oh, baby, I’m so good at listening to you flirt!"
#         )
#         i = i + 1


