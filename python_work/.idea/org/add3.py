import openpyxl
from openpyxl import load_workbook

def copy_data_with_repeats(source_path, target_path, repeat_times=2):
    # 加载源文件和目标文件
    source_wb = load_workbook(source_path)
    source_sheet = source_wb.active

    try:
        target_wb = load_workbook(target_path)
    except FileNotFoundError:
        # 如果目标文件不存在，创建一个新的
        target_wb = openpyxl.Workbook()

    target_sheet = target_wb.active

    # 设置目标表的表头
    target_sheet['A1'] = 'Orders'
    target_sheet['B1'] = 'Types'
    target_sheet['C1'] = 'Texts'

    target_row = 62  # 从第二行开始写入数据

    # 遍历源文件的每一行数据
    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        order, typ, text = row[0], row[1], row[2]

        # 将当前行数据重复指定次数写入目标文件
        for _ in range(repeat_times):
            target_sheet.cell(row=target_row, column=1, value=order)
            target_sheet.cell(row=target_row, column=2, value=typ)
            target_sheet.cell(row=target_row, column=3, value=text)
            target_row += 1

    # 保存目标文件
    target_wb.save(target_path)
    print(f"数据已成功复制并重复插入到 {target_path}")

# 使用示例
source_file = r'C:\Users\admin\Desktop\工作簿texts.xlsx'  # 替换为你的源文件路径
target_file = r'C:\Users\admin\Desktop\工作簿_1.xlsx'  # 替换为你的目标文件路径

copy_data_with_repeats(source_file, target_file)




# #前两行
# import openpyxl
# from openpyxl import load_workbook
#
# def copy_data_with_repeats(source_path, target_path, repeat_times=2):
#     # 加载源文件和目标文件
#     source_wb = load_workbook(source_path)
#     source_sheet = source_wb.active
#
#     try:
#         target_wb = load_workbook(target_path)
#     except FileNotFoundError:
#         # 如果目标文件不存在，创建一个新的
#         target_wb = openpyxl.Workbook()
#
#     target_sheet = target_wb.active
#
#     # 设置目标表的表头
#     target_sheet['A1'] = 'Orders'
#     target_sheet['B1'] = 'Types'
#
#     target_row = 2  # 从第二行开始写入数据
#
#     # 遍历源文件的每一行数据
#     for row in source_sheet.iter_rows(min_row=2, values_only=True):
#         order, typ = row[0], row[1]  # Changed to unpack only 2 values
#
#         # 将当前行数据重复指定次数写入目标文件
#         for _ in range(repeat_times):
#             target_sheet.cell(row=target_row, column=1, value=order)
#             target_sheet.cell(row=target_row, column=2, value=typ)
#             target_row += 1
#
#     # 保存目标文件
#     target_wb.save(target_path)
#     print(f"数据已成功复制并重复插入到 {target_path}")
#
# # 使用示例
# source_file = r'C:\Users\admin\Desktop\工作簿texts.xlsx'  # 替换为你的源文件路径
# target_file = r'C:\Users\admin\Desktop\工作簿_2.xlsx'  # 替换为你的目标文件路径
#
# copy_data_with_repeats(source_file, target_file)
